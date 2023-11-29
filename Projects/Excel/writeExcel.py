import openpyxl
from pathlib import Path

# create excel file
excelFile = openpyxl.Workbook()
print(excelFile.sheetnames)

# change sheet name
excelSheet =  excelFile.active
excelSheet.title = "firstSheet"

# create sheet
excelFile.create_sheet()
excelFile.create_sheet()
excelFile.create_sheet(index=1, title='secondSheet')
excelFile.create_sheet(index=2, title='middleSheet')

# delete sheet
del excelFile['middleSheet']

# write to sheet
sheet = excelFile['secondSheet']
sheet['A1'] = 'Hello, world!'
print(sheet['A1'].value)

# write to excel practice
names = ['Hadi', 'Yara', 'Sara', 'Anas']
firstSheet = excelFile['firstSheet']
for i in range(1, len(names)+1):
    firstSheet.cell(row=i, column=3).value = names[i-1]

# save excel file
excelFile.save(filename = Path.home() / Path('Desktop') / 'newExcel.xlsx')