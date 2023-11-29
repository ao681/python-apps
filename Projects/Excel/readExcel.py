import openpyxl
from pathlib import Path

excelFile = openpyxl.load_workbook(Path.home() / Path('Desktop','example.xlsx'))
print(excelFile.sheetnames)

Sheet1 = excelFile['Sheet1']
print(Sheet1.title)

activeSheet = excelFile.active
print(activeSheet.title)

print(Sheet1['A1'].value)
print(Sheet1['B1'].value)
print(Sheet1['C1'].value)
print(Sheet1['C1'].row)
print(Sheet1['C1'].column)
print(Sheet1['C1'].coordinate)

print(Sheet1.cell(row=1, column=2).value)

print('-' * 50)

for i in range(1, 7):
   print(i, Sheet1.cell(row=i, column=1).value)

print('-' * 50)

total = 0
for i in range(1, Sheet1.max_row+1):
   print(Sheet1.cell(row=i, column=1).value , Sheet1.cell(row=i, column=2).value)
   total += Sheet1.cell(row=i, column=2).value

print(f'The total salary of the employees is {total}')

print('-' * 50)
print(Sheet1.max_row)
print(Sheet1.max_column)